#!/usr/bin/python3
# -*- coding: UTF-8 -*-

import os
import time
from datetime import datetime
from openpyxl import Workbook, load_workbook, cell
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Border, Side, PatternFill, colors, Alignment

out_file_path = os.path.join(os.getcwd(), "../xlsx/sql语句.xlsx")

def remove_old_out_file(path):
    if os.path.exists(path):
        print("删除之前的输出文件")
        os.remove(path)

def get_all_excel_file_paths():
    print("开始查找excel文件(.xlsx)……")
    path = os.getcwd()
    file_paths = []
    print(path)
    for file_name in os.listdir(path):
        print(file_name)
        if file_name.endswith("xlsx") and not (file_name.startswith("~") or file_name.startswith(".")):
            print("找到excel文件：", file_name)
            file_paths.append(os.path.join(path, file_name))
    if len(file_paths) == 0:
        print("没有找到excel文件")

    if len(file_paths) > 1:
        print("目前只支持转1个excel文件")
        exit()
    return file_paths


def get_custom_titles_map_from_sheet(sheet):
    custom_title_map = {}
    for row in sheet.iter_rows(min_row=1, max_row=1, min_col=1, max_col=sheet.max_column):
        for cell in row:
            custom_title = cell.value
            if custom_title is not None:
                custom_title_map[custom_title] = cell.column
    return custom_title_map


def merge_sheet(source_sheet, target_sheet):
    custom_titles_map = get_custom_titles_map_from_sheet(target_sheet)
    target_sheet_row_count = target_sheet.max_row

    model_file_path = source_sheet.title+".txt"
    if not os.path.exists(model_file_path):
        return

    model_file = open(model_file_path, 'r')
    sql = model_file.read()
    for index, rows in enumerate(source_sheet.rows):
        if index == 0:
            continue
        new_sql = sql
        for cell in rows:
            if cell.value is not None:
                key = "cell"+str(cell.column)
                new_sql = new_sql.replace(key, cell.value, 1)
            else:
                print("空内容  位置：列："+cell.column + "-行：" + cell.row)
                exit()

        new_cell = target_sheet.cell(row=index,
                                     column=1,
                                     value=new_sql)
    model_file.close()

def start_merge():
    remove_old_out_file(out_file_path)
    new_workbook = Workbook()

    in_file_paths = get_all_excel_file_paths()
    for file_path in in_file_paths:
        print("开始处理文件:", os.path.basename(file_path))
        workbook = load_workbook(file_path)
        for sheet in workbook:
            sheetname = sheet.title
            if sheetname not in new_workbook.sheetnames:
                new_workbook.create_sheet(title=sheetname)
            merge_sheet(workbook[sheetname], new_workbook[sheetname])
            print("sheet 处理完成:", sheetname)
        workbook.close()
        print("文件处理完成:", os.path.basename(file_path))

    new_workbook.save(out_file_path)
    new_workbook.close()


if __name__ == '__main__':
    print("开始执行任务……")
    start_time = datetime.now()
    start_merge()
    delta = (datetime.now() - start_time).seconds
    print("任务执行完毕, 共耗时:", int(delta / 60), "分", delta % 60, "秒")


