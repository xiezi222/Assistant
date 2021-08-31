#!/usr/bin/python3
# -*- coding: UTF-8 -*-

import os
from base_script import BaseScript
from openpyxl import Workbook, load_workbook


class ExcelToSql(BaseScript):

    def __init__(self, excel_path, sql_path):
        self.excel_path = excel_path
        self.sql_path = sql_path
        self.out_path = os.path.join(os.path.dirname(excel_path), "sql语句.xlsx")

    def remove_old_out_file(self):
        if os.path.exists(self.out_path):
            print("删除之前的输出文件")
            os.remove(self.out_path)

    def get_all_excel_file_paths(self):
        print("开始查找excel文件(.xlsx)……")
        path = os.getcwd()
        file_paths = []
        print(path)
        for file_name in os.listdir(path):
            print(file_name)
            if file_name.endswith("xlsx") and not (file_name.startswith("~") or file_name.startswith(".")):
                print("找到excel文件：", file_name)
                file_paths.append(os.path.join(path, file_name))
        if len(file_paths) == 0:
            print("没有找到excel文件")

        if len(file_paths) > 1:
            print("目前只支持转1个excel文件")
            exit()
        return file_paths

    def get_custom_titles_map_from_sheet(self, sheet):
        custom_title_map = {}
        for row in sheet.iter_rows(min_row=1, max_row=1, min_col=1, max_col=sheet.max_column):
            for cell in row:
                custom_title = cell.value
                if custom_title is not None:
                    custom_title_map[custom_title] = cell.column
        return custom_title_map

    def merge_sheet(self, source_sheet, target_sheet):
        custom_titles_map = self.get_custom_titles_map_from_sheet(target_sheet)
        target_sheet_row_count = target_sheet.max_row

        model_file_path = os.path.join(self.sql_path, source_sheet.title+".txt")
        if not os.path.exists(model_file_path):
            return

        model_file = open(model_file_path, 'r')
        sql = model_file.read()
        for index, rows in enumerate(source_sheet.rows):
            if index == 0:
                continue
            new_sql = sql
            for cell in rows:
                if cell.value is not None:
                    key = "cell" + str(cell.column)
                    new_sql = new_sql.replace(key, cell.value, 1)
                else:
                    print("空内容  位置：列：" + cell.column + "-行：" + cell.row)
                    exit()

            new_cell = target_sheet.cell(row=index,
                                         column=1,
                                         value=new_sql)
        model_file.close()

    def executive(self):
        self.remove_old_out_file()
        new_workbook = Workbook()

        print("开始处理文件:", self.excel_path)
        workbook = load_workbook(self.excel_path)
        for sheet in workbook:
            sheetname = sheet.title
            if sheetname not in new_workbook.sheetnames:
                new_workbook.create_sheet(title=sheetname)
            self.merge_sheet(workbook[sheetname], new_workbook[sheetname])
            print("sheet 处理完成:", sheetname)
        workbook.close()
        print("文件处理完成:", self.excel_path)

        new_workbook.save(self.out_path)
        new_workbook.close()



